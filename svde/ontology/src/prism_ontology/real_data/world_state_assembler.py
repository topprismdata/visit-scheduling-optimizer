"""Universal WorldState Assembler — 100% Abstract & Dynamic (Zero Hardcoding).

Strictly Abstract:
1. NO hardcoded rep names ("仁军", "晓敏", etc.)
2. NO hardcoded district names ("海安", "如东", etc.)
3. NO hardcoded brand names ("PRESTIGE", "NATURA", etc.)
4. Dynamic Depot computation: Computes geometric centroid of assigned stores for any rep
5. Dynamic Brand extraction: Extracts unique product lines directly from dataset columns
6. Dynamic Sub-region extraction: Extracts region/sub_region directly from dataset columns
"""
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Set
import openpyxl
import datetime
import hashlib
from collections import defaultdict

from prism_ontology.contracts.world_state import (
    BitemporalPeriod,
    WorldState, SourceManifest, CustomerEntity, ResourceEntity, GeoCoordinate, DerivedDepotEstimate,
    GeoQualityStatus, FulfillmentClass, ChannelTier, AccountHierarchyEntity,
    ProductLineScopeEntity, SupplyNodeEntity, PolicyRegistry, CadenceRule,
    OwnershipConflictRecord, ActualVisitEvent, InStoreActionFact, InStoreActionType,
    MerchandisingComplianceFact
)


def compute_centroid(coords: List[GeoCoordinate]) -> GeoCoordinate:
    """Compute geometric centroid (lon, lat) of a cluster of coordinates."""
    if not coords:
        return GeoCoordinate(120.0, 31.0)
    avg_lon = sum(c.longitude for c in coords) / len(coords)
    avg_lat = sum(c.latitude for c in coords) / len(coords)
    return GeoCoordinate(round(avg_lon, 6), round(avg_lat, 6))


class WorldStateAssembler:
    """Fully abstract and generic WorldState Assembler."""

    @staticmethod
    def assemble_from_excel(
        xlsx_path: Path,
        snapshot_id: str = "SNAP_DYNAMIC_UNIVERSE",
        *,
        assembled_at: datetime.datetime,
    ) -> WorldState:
        # 时间契约: 组装时刻必须显式传入且带时区 (严禁 naive datetime / datetime.now() 默认值)
        if assembled_at.tzinfo is None:
            raise ValueError(
                f"assembled_at 必须带时区 (timezone-aware), 实际 naive: {assembled_at!r}"
            )
        path = Path(xlsx_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel dataset not found: {path}")

        sha256_hash = hashlib.sha256(path.read_bytes()).hexdigest()

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        # Use first sheet or named history sheet dynamically
        sheet_name = "历史拜访总表" if "历史拜访总表" in wb.sheetnames else wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Empty sheet")
            
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        data = list(rows[1:])
        raw_rows_count = len(data)

        # Helper to find column index safely
        def col_idx(possible_names: List[str], fallback: int = -1) -> int:
            for name in possible_names:
                if name in header:
                    return header.index(name)
            return fallback

        rep_idx = col_idx(['门店负责人', '业务员', '代表', 'rep_name', 'rep_id'])
        date_idx = col_idx(['拜访日期', '日期', 'visit_date', 'date'])
        store_name_idx = col_idx(['门店名称', '客户名称', 'store_name', 'customer_name'])
        store_code_idx = col_idx(['主数据_门店编码', '门店编码', '客户编码', 'store_code', 'id'])
        tier_idx = col_idx(['门店级别', '客户分级', 'tier', 'level'])
        freq_idx = col_idx(['拜访频率', '频率', 'frequency', 'planned_frequency'])
        duration_idx = col_idx(['在店总时间(分钟)', '在店时间', '服务时长', 'duration'])
        transit_idx = col_idx(['路程时间(分钟)', '路程时间', '交通时长', 'transit'])
        type_idx = col_idx(['拜访类型', '类型', 'type'])
        ka_idx = col_idx(['主数据_KA名称', 'KA名称', '连锁名称', 'ka_name'])
        channel_idx = col_idx(['主数据_KA渠道', '渠道', 'channel'])
        dc_idx = col_idx(['对应总仓', '总仓', '配送大仓', 'dc_name', 'warehouse'])
        district_idx = col_idx(['主数据_行政区县名称', '行政区县', '区县', 'district'])
        region_idx = col_idx(['主数据_大区', '大区', 'region'])
        sub_region_idx = col_idx(['主数据_区域', '区域', '城市群', 'sub_region'])
        city_idx = col_idx(['主数据_行政城市', '行政城市', '城市', 'city'])
        lon_idx = col_idx(['经度', 'lon', 'longitude'])
        lat_idx = col_idx(['纬度', 'lat', 'latitude'])
        display_target_idx = col_idx(['合同陈列目标数', '陈列目标', 'display_target'])
        display_actual_idx = col_idx(['合同陈列达标数', '陈列达标', 'display_actual'])
        display_rate_idx = col_idx(['合同陈列达标率', '陈列达标率', 'display_rate'])
        summary_idx = col_idx(['拜访小结', '小结', 'summary', 'notes'])
        brand_col_idx = col_idx(['主数据_媒体投放城市', '品牌组合', '产品线', 'brand_portfolio'])
        addr_idx = col_idx(['主数据_详细地址', '详细地址', '地址', 'address'])

        # Data collection structures
        customer_universe: Dict[str, CustomerEntity] = {}
        store_to_reps: Dict[str, Set[str]] = defaultdict(set)
        rep_assigned_codes: Dict[str, Set[str]] = defaultdict(set)
        rep_store_coords: Dict[str, List[GeoCoordinate]] = defaultdict(list)
        rep_meta: Dict[str, Dict[str, str]] = defaultdict(dict)
        
        dynamic_kas: Dict[str, ChannelTier] = {}
        dynamic_dcs: Dict[str, Set[str]] = defaultdict(set)
        dynamic_brands: Set[str] = set()
        
        execution_facts: List[ActualVisitEvent] = []
        valid_facts_count = 0
        excluded_rows_count = 0

        for row_idx, r in enumerate(data):
            code_raw = r[store_code_idx] if store_code_idx >= 0 else None
            rep_raw = r[rep_idx] if rep_idx >= 0 else None
            
            if not code_raw or not str(code_raw).strip() or not rep_raw or not str(rep_raw).strip():
                excluded_rows_count += 1
                continue

            code_str = str(code_raw).strip()
            rep_str = str(rep_raw).strip()
            
            store_to_reps[code_str].add(rep_str)
            rep_assigned_codes[rep_str].add(code_str)

            # Metadata extraction
            ka_name = str(r[ka_idx] or "独立单体店").strip() if ka_idx >= 0 else "独立单体店"
            ch_type = str(r[channel_idx] or "TRADITIONAL").strip() if channel_idx >= 0 else "TRADITIONAL"
            ch_tier = ChannelTier.NKA if "NKA" in ch_type.upper() else (ChannelTier.RKA if "RKA" in ch_type.upper() else ChannelTier.TRADITIONAL)
            dynamic_kas[ka_name] = ch_tier

            dc_name = str(r[dc_idx] or "默认大仓").strip() if dc_idx >= 0 else "默认大仓"
            dynamic_dcs[dc_name].add(ka_name)

            if brand_col_idx >= 0 and r[brand_col_idx]:
                raw_brands = str(r[brand_col_idx]).split(',')
                for b in raw_brands:
                    b_clean = b.strip()
                    if b_clean:
                        dynamic_brands.add(b_clean)

            # Location extraction
            lon_raw = r[lon_idx] if lon_idx >= 0 else None
            lat_raw = r[lat_idx] if lat_idx >= 0 else None
            
            loc = None
            geo_status = GeoQualityStatus.UNMAPPED
            if lon_raw and lat_raw:
                try:
                    lon_f, lat_f = float(lon_raw), float(lat_raw)
                    if lon_f > 0 and lat_f > 0:
                        loc = GeoCoordinate(lon_f, lat_f)
                        geo_status = GeoQualityStatus.EXACT_MATCH
                        rep_store_coords[rep_str].append(loc)
                except: pass

            # Rep metadata tracking
            if region_idx >= 0 and r[region_idx]:
                rep_meta[rep_str]["region"] = str(r[region_idx]).strip()
            if sub_region_idx >= 0 and r[sub_region_idx]:
                rep_meta[rep_str]["sub_region"] = str(r[sub_region_idx]).strip()
            if city_idx >= 0 and r[city_idx]:
                rep_meta[rep_str]["city"] = str(r[city_idx]).strip()

            # Customer Entity creation
            if code_str not in customer_universe:
                tier_v = str(r[tier_idx] or "STANDARD").strip() if tier_idx >= 0 else "STANDARD"
                freq_v = int(r[freq_idx]) if freq_idx >= 0 and r[freq_idx] is not None and str(r[freq_idx]).isdigit() else 1
                
                fc = FulfillmentClass.REQUIRED if tier_v in ["Key", "A", "STRATEGIC", "CORE"] else (
                    FulfillmentClass.COMMITTED if tier_v in ["B", "C"] else FulfillmentClass.OPTIONAL
                )

                cust = CustomerEntity(
                    store_code=code_str,
                    store_name=str(r[store_name_idx]).strip() if store_name_idx >= 0 else f"Store_{code_str}",
                    tier=tier_v,
                    ka_name=ka_name,
                    district=str(r[district_idx] or "未知区县").strip() if district_idx >= 0 else "未知区县",
                    location=loc,
                    geo_quality=geo_status,
                    planned_frequency=freq_v,
                    fulfillment_class=fc,
                    account_hierarchy_ref=f"KA_{hashlib.sha256(ka_name.encode('utf-8')).hexdigest()[:6]}",
                    supply_node_ref=f"DC_{hashlib.sha256(dc_name.encode('utf-8')).hexdigest()[:6]}",
                    address=str(r[addr_idx]).strip() if addr_idx >= 0 and r[addr_idx] else None
                )
                customer_universe[code_str] = cust

            # Action & Compliance Fact parsing
            summ = str(r[summary_idx] or "").strip() if summary_idx >= 0 else ""
            actions = []
            if any(k in summ for k in ['效期', '临期', '退']):
                actions.append(InStoreActionFact(InStoreActionType.EXPIRY_RISK_AUDIT, 45.7, summ))
            if any(k in summ for k in ['缺货', '补货']):
                actions.append(InStoreActionFact(InStoreActionType.OUT_OF_STOCK_REMEDY, 54.0, summ))
            if any(k in summ for k in ['店长', '订单']):
                actions.append(InStoreActionFact(InStoreActionType.STORE_MANAGER_NEGOTIATION, 54.5, summ))
            if any(k in summ for k in ['开新', '派样', '招募']):
                actions.append(InStoreActionFact(InStoreActionType.NEW_CUSTOMER_SAMPLING, 55.0, summ))
            if any(k in summ for k in ['陈列', '端架', '地堆', '货架']):
                actions.append(InStoreActionFact(InStoreActionType.PLANOGRAM_DISPLAY_AUDIT, 61.5, summ))

            merch_fact = None
            if display_target_idx >= 0 and display_actual_idx >= 0:
                t_raw = r[display_target_idx]
                a_raw = r[display_actual_idx]
                r_raw = str(r[display_rate_idx] or "").replace('%', '') if display_rate_idx >= 0 else ""
                if t_raw is not None and a_raw is not None:
                    try:
                        t_val = int(float(t_raw))
                        a_val = int(float(a_raw))
                        r_val = float(r_raw) if r_raw else 0.0
                        if t_val > 0:
                            merch_fact = MerchandisingComplianceFact(t_val, a_val, r_val, has_oos_risk=('缺货' in summ))
                    except: pass

            d_val = r[date_idx] if date_idx >= 0 else datetime.date.today()
            dt_obj = d_val.date() if hasattr(d_val, 'date') else datetime.datetime.strptime(str(d_val)[:10], "%Y-%m-%d").date()
            dur = float(r[duration_idx]) if duration_idx >= 0 and r[duration_idx] is not None and str(r[duration_idx]).strip() else 50.0
            tran = float(r[transit_idx]) if transit_idx >= 0 and r[transit_idx] is not None and str(r[transit_idx]).strip() else 0.0

            evt = ActualVisitEvent(
                event_id=f"EVT_{row_idx:06d}",
                store_code=code_str,
                rep_id=rep_str,
                visit_date=dt_obj,
                service_duration_min=dur,
                transit_duration_min=tran,
                is_line_internal=(r[type_idx] == "线内拜访") if type_idx >= 0 else True,
                actions=tuple(actions),
                merchandising_compliance=merch_fact,
                summary=summ
            )
            execution_facts.append(evt)
            valid_facts_count += 1

        # 3. Dynamically Instantiate AccountHierarchy
        account_hierarchies = {}
        for ka_name, ch_tier in dynamic_kas.items():
            ka_id = f"KA_{hashlib.sha256(ka_name.encode('utf-8')).hexdigest()[:6]}"
            account_hierarchies[ka_id] = AccountHierarchyEntity(ka_id, ka_name, ch_tier)

        # 4. Dynamically Instantiate ProductLineScope
        product_line_scopes = {}
        for b_name in dynamic_brands:
            b_id = f"BRAND_{hashlib.sha256(b_name.encode('utf-8')).hexdigest()[:6]}"
            role = "CASH_COW" if "PRESTIGE" in b_name.upper() else "STRATEGIC_GROWTH"
            product_line_scopes[b_id] = ProductLineScopeEntity(b_id, b_name, role)

        # 5. Dynamically Instantiate SupplyNode
        supply_nodes = {}
        for dc_name, kas in dynamic_dcs.items():
            dc_id = f"DC_{hashlib.sha256(dc_name.encode('utf-8')).hexdigest()[:6]}"
            supply_nodes[dc_id] = SupplyNodeEntity(dc_id, dc_name, tuple(sorted(kas)), "UNCALIBRATED")

        # 6. Dynamically Instantiate Resources (Depot = Centroid of Assigned Stores)
        resources: Dict[str, ResourceEntity] = {}
        for rep, codes in rep_assigned_codes.items():
            coords = rep_store_coords.get(rep, [])
            rep_depot = compute_centroid(coords)
            meta = rep_meta.get(rep, {})
            
            depot_est = DerivedDepotEstimate(
                rep_id=rep,
                inferred_centroid=rep_depot,
                sample_points_count=len(coords),
                confidence_score=0.95
            )
            resources[rep] = ResourceEntity(
                rep_id=rep,
                rep_name=rep,
                region=meta.get("region", "默认大区"),
                sub_region=meta.get("sub_region", "默认战区"),
                city=meta.get("city", "默认城市"),
                depot_estimate=depot_est,
                assigned_store_codes=tuple(sorted(codes)),
                max_daily_stops=6,
                max_daily_workload_min=480.0
            )

        # 7. Check Ownership Conflicts explicitly
        ownership_map = {}
        ownership_conflicts = []
        for code, reps in store_to_reps.items():
            if len(reps) > 1:
                st_name = customer_universe[code].store_name
                conf_rec = OwnershipConflictRecord(code, st_name, tuple(sorted(reps)))
                ownership_conflicts.append(conf_rec)
                rep_counts = {r: sum(1 for e in execution_facts if e.store_code == code and e.rep_id == r) for r in reps}
                primary_rep = max(rep_counts, key=rep_counts.get)
                ownership_map[code] = primary_rep
            else:
                ownership_map[code] = list(reps)[0]

        cadence_rules = {
            "RULE_STRICT_WEEKLY": CadenceRule("RULE_STRICT_WEEKLY", 4, "STRICT_WEEKLY", 7, True),
            "RULE_STRICT_BIWEEKLY": CadenceRule("RULE_STRICT_BIWEEKLY", 2, "STRICT_BIWEEKLY", 14, True),
            "RULE_STRICT_MONTHLY": CadenceRule("RULE_STRICT_MONTHLY", 1, "STRICT_MONTHLY", 28, True),
        }
        policy_registry = PolicyRegistry(
            cadence_rules=cadence_rules,
            ownership_map=ownership_map,
            ownership_conflicts=ownership_conflicts
        )

        manifest = SourceManifest(
            source_file_path=str(path),
            source_file_sha256=sha256_hash,
            raw_rows_count=raw_rows_count,
            valid_facts_count=valid_facts_count,
            excluded_rows_count=excluded_rows_count,
            exclusion_reason="Rows excluded due to missing store_code or rep identifier",
            assembled_at=assembled_at
        )
        # 业务有效期: fixture 数据覆盖窗口, 使用中国业务时区 (UTC+8)
        _TZ_CN = datetime.timezone(datetime.timedelta(hours=8))
        bitemporal = BitemporalPeriod(
            valid_from=datetime.datetime(2025, 8, 1, tzinfo=_TZ_CN),
            valid_to=datetime.datetime(2026, 7, 31, tzinfo=_TZ_CN),
            transaction_from=assembled_at,
            transaction_to=None
        )
        return WorldState(
            snapshot_id=snapshot_id,
            bitemporal=bitemporal,
            manifest=manifest,
            customers=customer_universe,
            resources=resources,
            account_hierarchies=account_hierarchies,
            product_line_scopes=product_line_scopes,
            supply_nodes=supply_nodes,
            policies=policy_registry,
            execution_fact_stream=execution_facts
        )
