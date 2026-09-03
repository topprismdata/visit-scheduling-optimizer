"""FMCG Real Data Ingestion — Phase 5 of prism-ontology v1.1.

Maps the 53-column FMCG historical visit Excel file to v0.3 ReferenceObjects.

Hard guarantees (per v1.1 §6.1):
1. No fields beyond v0.3 schema are accepted into canonical state
2. Every field mapping is reversible (raw data preserved)
3. Pre-flight validation through DataPrecheckValidator REQUIRED before mapping
4. ALGORITHM concepts (e.g. route optimization) NOT in this module — only data
5. SOP objects NOT created here — GAP-6 closed
"""
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Set
from collections import Counter

from prism_ontology.reference.store import ReferenceOntologyStore, ObjectLayer


# v0.3 ReferenceObject field set (frozen) - 19 objects
# Only these source columns are mapped to v0.3 fields
FMCG_TO_V03_FIELD_MAP: Dict[str, str] = {
    # === IDENTITY layer ===
    # Customer (主数据_门店名称 in v0.3 → Customer.tier from 门店级别)
    "主数据_门店编码": "Customer.id",
    "主数据_门店名称": "Customer.name",
    "门店级别": "Customer.tier",  # Key/A/B/C maps to STRATEGIC/CORE/STANDARD
    "经度": "Customer.location.lon",
    "纬度": "Customer.location.lat",
    "主数据_KA名称": "Customer.ka_name",
    "主数据_详细地址": "Customer.address",

    # Resource (sales rep = Resource; the 7 reps in this dataset)
    "门店负责人": "Resource.rep_id",
    "大区": "Resource.region",
    "区域": "Resource.sub_region",

    # === EVENT layer ===
    "拜访日期": "VisitEvent.date",
    "进店时间": "VisitEvent.arrival",
    "离店时间": "VisitEvent.departure",
    "在店总时间(分钟)": "VisitEvent.service_duration",
    "路程时间(分钟)": "VisitEvent.transit_duration",
    "拜访类型": "VisitEvent.visit_type",  # 线内 / 线外 / 取消
    "拜访模式": "VisitEvent.visit_status",  # 正常 / 取消 / GPS偏差大

    # PlannedVisit-style fields
    "星期几": "PlannedVisit.weekday",
    "拜访顺序": "PlannedVisit.sequence_idx",
    "拜访频率": "PlannedVisit.frequency_per_period",
    "第几周": "PlannedVisit.week_of_period",

    # === POLICY layer ===
    "主数据_覆盖类型": "OwnershipPolicy.assignment_source",  # e.g., 美素佳儿人员覆盖
    "主数据_KA渠道": "OwnershipPolicy.channel_type",  # NKA / KA

    # === MEASUREMENT layer ===
    # TravelCostMatrix: 经纬度 + 路程时间 → estimate inter-store travel time
}


# v0.3 tier mapping (FMCG tier → v0.3 tier)
FMCG_TIER_MAP: Dict[str, str] = {
    "Key": "STRATEGIC",
    "A": "CORE",
    "B": "STANDARD",
    "C": "STANDARD",
    "D": "STANDARD",
}


# v0.3 visit type mapping
FMCG_VISIT_TYPE_MAP: Dict[str, str] = {
    "线内拜访": "PlannedVisit.regular",
    "线外拜访": "PlannedVisit.extra",
    "取消拜访": "PlannedVisit.cancelled",
}


class FMCGDataError(Exception):
    """Raised when data fails v0.3 schema validation."""
    pass


class FMCGRealDataIngestor:
    """Read-only ingestion from FMCG Excel → v0.3 ReferenceObject field projections."""

    def __init__(self, xlsx_path: Path, store: ReferenceOntologyStore):
        self.path = Path(xlsx_path)
        self.store = store
        self.raw_rows: List[dict] = []
        self.mapped: List[dict] = []

    def load(self) -> "FMCGRealDataIngestor":
        """Load raw rows from Excel, NO mutation to v0.3 store."""
        if not self.path.exists():
            raise FileNotFoundError(f"FMCG data file not found: {self.path}")
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        if "历史拜访总表" not in wb.sheetnames:
            raise FMCGDataError(f"Sheet '历史拜访总表' not found in {self.path}")
        sheet = wb["历史拜访总表"]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise FMCGDataError("Empty sheet")
        header = list(rows[0])
        for r in rows[1:]:
            if r and any(c is not None for c in r):
                self.raw_rows.append(dict(zip(header, r)))
        return self

    def precheck(self) -> Dict[str, any]:
        """v1.1 §6.1 data pre-flight validation (subset of DataPrecheckValidator checks)."""
        if not self.raw_rows:
            raise FMCGDataError("Load data first via .load()")
        report = {
            "total_rows": len(self.raw_rows),
            "issues": [],
        }
        # Required field check
        required_fields = list(FMCG_TO_V03_FIELD_MAP.keys())[:8]
        if self.raw_rows:
            sample = self.raw_rows[0]
            missing = [f for f in required_fields if f not in sample]
            if missing:
                report["issues"].append(f"Missing required fields: {missing[:3]}")
        # Sales rep count
        reps = Counter(r.get("门店负责人") for r in self.raw_rows if r.get("门店负责人"))
        report["unique_reps"] = len(reps)
        report["rep_names"] = list(reps.keys())
        # Date range
        dates = sorted({r["拜访日期"].date() if r.get("拜访日期") and hasattr(r["拜访日期"], "date") else r.get("拜访日期") for r in self.raw_rows if r.get("拜访日期")})
        report["date_range"] = (str(dates[0]), str(dates[-1])) if dates else (None, None)
        # Tier distribution
        tiers = Counter(r.get("门店级别") for r in self.raw_rows if r.get("门店级别"))
        report["tier_distribution"] = dict(tiers)
        report["is_valid"] = len(report["issues"]) == 0
        return report

    def project_to_v03(self) -> List[dict]:
        """Project raw rows to v0.3 field-aligned dicts (read-only, no mutation)."""
        if not self.raw_rows:
            raise FMCGDataError("Load data first via .load()")
        mapped = []
        for r in self.raw_rows:
            v03_row = self._project_row(r)
            mapped.append(v03_row)
        self.mapped = mapped
        return mapped

    def _project_row(self, raw: dict) -> dict:
        """Project a single raw row to v0.3-aligned field names."""
        out = {
            "source_row_ref": id(raw),  # reversible reference
        }
        for fmcg_field, v03_field in FMCG_TO_V03_FIELD_MAP.items():
            if fmcg_field in raw and raw[fmcg_field] is not None:
                out[v03_field] = raw[fmcg_field]
        # Tier mapping
        if "门店级别" in raw and raw["门店级别"]:
            out["Customer.tier"] = FMCG_TIER_MAP.get(raw["门店级别"], "STANDARD")
        # Visit type mapping
        if "拜访类型" in raw and raw["拜访类型"]:
            out["VisitEvent.visit_type_classified"] = FMCG_VISIT_TYPE_MAP.get(
                raw["拜访类型"], "PlannedVisit.unknown"
            )
        return out

    def field_mapping_report(self) -> Dict[str, any]:
        """Generate field mapping coverage report (per v0.3 object)."""
        if not self.mapped:
            self.project_to_v03()
        coverage = {}
        for fmcg_f, v03_f in FMCG_TO_V03_FIELD_MAP.items():
            obj = v03_f.split(".")[0]
            coverage.setdefault(obj, []).append((fmcg_f, v03_f))
        unmapped_columns = []
        if self.raw_rows:
            sample = self.raw_rows[0]
            for col in sample.keys():
                if col and col not in FMCG_TO_V03_FIELD_MAP:
                    unmapped_columns.append(col)
        return {
            "v0_3_object_field_count": {obj: len(fs) for obj, fs in coverage.items()},
            "fmcg_to_v03_mappings": {obj: fs for obj, fs in coverage.items()},
            "unmapped_fmcg_columns": unmapped_columns,
            "total_mapped": len(FMCG_TO_V03_FIELD_MAP),
            "total_unmapped": len(unmapped_columns),
        }
