"""SFA 进离店打卡报表 → 规范 OperationalDecisionWorldState 适配器 + 数据可信度清洗.

输入: 移动端 SFA 导出的「进离店报表」xlsx (列: 大区/办事处/地区/辖区/片区/人员编码/
人员名称/客户编码/客户名称/进店时间/离店时间/在店时长(分钟)/进店经度/进店纬度/
偏差(米)/打卡状态/自动离店 ...)

与 WorldStateAssembler 的区别:
- assembler 面向「历史拜访总表」(一行一拜访, 含拜访频率/小结);
- 本模块面向「进离店打卡流水」(原始执行事实流, 无计划字段, 含 GPS 可信度信号).

四条数据清洗规则 (参数化, 输出 audit trail; 观察事实绝不静默删除):
- R1 CHECKOUT_INFLATION   在店时长 >= 上限或自动离店 -> service_duration 截断
- R2 BATCH_CHECKIN_SUSPECT 连续同坐标 + 间隔<=阈值 + 在店<=阈值 -> 降权 (credit=0)
- R3 GPS_DEVIANCE         偏差(米) > 阈值 -> 位置不可信 (不删事件, 打信用标记)
- R4 STORE_COORD_DRIFT    同一客户可信坐标跨度 > 阈值 -> 主数据位置改聚类质心

降权/截断信息写入 ActualVisitEvent.summary 与 InStoreActionFact.action_notes (可审计)。
每客户有效月度频次观测作为 OperationalVisitPolicy 写入 policy registry
(approved_by=DERIVED_FROM_OBSERVATION — 显式标注派生, 绝不冒充签署政策)。
"""
from __future__ import annotations

import datetime
import hashlib
import math
from collections import defaultdict
from dataclasses import dataclass, field, replace
from pathlib import Path
from statistics import mean
from typing import Dict, List, Optional, Set, Tuple

import openpyxl

from prism_ontology.world_model.state_snapshot import (
    BitemporalPeriod,
    CadenceRule,
    ChannelTier,
    DerivedDepotEstimate,
    FulfillmentClass,
    GeoCoordinate,
    GeoQualityStatus,
    InStoreActionFact,
    InStoreActionType,
    OperationalCustomer,
    OperationalDecisionWorldState,
    OperationalResource,
    OperationalVisitPolicy,
    ActualVisitEvent,
    AccountHierarchyEntity,
    PolicyRegistry,
    SourceManifest,
)

TZ_CN = datetime.timezone(datetime.timedelta(hours=8))


# ---------------------------------------------------------------- 清洗参数

@dataclass(frozen=True)
class CleaningParams:
    r1_cap_min: int = 120            # R1: 在店时长截断上限 (分钟)
    r2_gap_max_min: float = 5.0      # R2: 与前次离店的最大间隔 (分钟)
    r2_short_dur_max: int = 2        # R2: 在店 <= N 分钟视为秒打卡
    r2_same_coord_eps: float = 1e-6  # R2: 经纬度一致判定容差
    r3_deviation_max_m: float = 100  # R3: GPS 偏差阈值 (米)
    r4_drift_deg: float = 0.01       # R4: 客户坐标跨度阈值 (~1km)


@dataclass
class CleaningStats:
    raw_events: int = 0
    r1_truncated: int = 0
    r2_suspects: int = 0
    r3_gps_bad: int = 0
    r4_drift_customers: int = 0
    r4_total_customers: int = 0
    unmapped_customers: int = 0
    effective_events: int = 0        # 未降权事件数 (credit > 0)

    @property
    def reliability_rate(self) -> float:
        return self.effective_events / self.raw_events if self.raw_events else 0.0


@dataclass(frozen=True)
class EventFlag:
    code: str      # R1/R2/R3/R4 规则码
    target: str    # EVT:<cust>@<time> | CUST:<cust>
    note: str


# ---------------------------------------------------------------- 工具

def haversine_m(a: GeoCoordinate, b: GeoCoordinate) -> float:
    R = 6371000.0
    p1, p2 = math.radians(a.latitude), math.radians(b.latitude)
    dp = p2 - p1
    dl = math.radians(b.longitude - a.longitude)
    h = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


def _parse_dt(v) -> Optional[datetime.datetime]:
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, str) and v.strip():
        return datetime.datetime.strptime(v.strip()[:19], "%Y-%m-%d %H:%M:%S")
    return None


def _to_float(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _coord_cluster(coords: List[GeoCoordinate], eps_deg: float) -> Tuple[List[GeoCoordinate], int]:
    """增量空间聚类: 返回 (最大簇成员, 最大簇大小)."""
    reps: List[List] = []  # [centroid, count]
    for c in coords:
        placed = False
        for slot in reps:
            rc, n = slot[0], slot[1]
            if abs(c.longitude - rc.longitude) <= eps_deg and abs(c.latitude - rc.latitude) <= eps_deg:
                slot[0] = GeoCoordinate(round((rc.longitude * n + c.longitude) / (n + 1), 6),
                                        round((rc.latitude * n + c.latitude) / (n + 1), 6))
                slot[1] = n + 1
                placed = True
                break
        if not placed:
            reps.append([c, 1])
    if not reps:
        return [], 0
    big = max(reps, key=lambda s: s[1])
    members = [c for c in coords
               if abs(c.longitude - big[0].longitude) <= eps_deg
               and abs(c.latitude - big[0].latitude) <= eps_deg]
    return members, big[1]


# ---------------------------------------------------------------- 主适配器

class SFACheckinIngestor:
    """进离店打卡报表 -> 规范 WorldState (只读源文件; 清洗只影响内存模型)."""

    REQUIRED_COLS = ["人员编码", "人员名称", "客户编码", "客户名称",
                     "进店时间", "离店时间", "在店时长(分钟)",
                     "进店经度", "进店纬度", "偏差(米)"]

    @classmethod
    def assemble_from_excel(
        cls,
        xlsx_path: str,
        *,
        assembled_at: datetime.datetime,
        params: CleaningParams = CleaningParams(),
        snapshot_id: str = "SNAP_SFA_CHECKIN",
    ) -> Tuple[OperationalDecisionWorldState, List[EventFlag], CleaningStats]:
        """解析 + 清洗 + 组装. 返回 (WorldState, audit flags, stats)."""
        if assembled_at.tzinfo is None:
            raise ValueError(f"assembled_at 必须带时区 (timezone-aware), 实际 naive: {assembled_at!r}")
        path = Path(xlsx_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel not found: {path}")
        sha256 = hashlib.sha256(path.read_bytes()).hexdigest()
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        # SFA 导出文件常缺 <dimension> 元数据; read_only 模式下必须重置维度,
        # 否则 iter_rows 只吐第一格 (openpyxl 已知行为)
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        rows = list(sheet.iter_rows(values_only=True))
        wb.close()
        if not rows:
            raise ValueError("Empty sheet")
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        missing = [c for c in cls.REQUIRED_COLS if c not in header]
        if missing:
            raise ValueError(f"缺少必需列: {missing}")
        ix = {name: header.index(name) for name in header}

        def cell(r, name: str, default=""):
            i = ix.get(name, -1)
            if i < 0 or r[i] is None:
                return default
            return str(r[i]).strip()

        # ---- Pass 1: 行 -> 原始事件 ----
        raw: List[dict] = []
        excluded = 0
        for r_i, r in enumerate(rows[1:]):
            code = cell(r, "客户编码")
            rep = cell(r, "人员名称")
            t_in = _parse_dt(r[ix["进店时间"]])
            t_out = _parse_dt(r[ix["离店时间"]])
            if not code or not rep or t_in is None or t_out is None:
                excluded += 1
                continue
            raw.append(dict(
                row=r_i, code=code, rep=rep,
                name=cell(r, "客户名称", code),
                t_in=t_in, t_out=t_out,
                dur=_to_float(r[ix["在店时长(分钟)"]]),
                lon=_to_float(r[ix["进店经度"]]), lat=_to_float(r[ix["进店纬度"]]),
                dev=_to_float(r[ix["偏差(米)"]]),
                org=cell(r, "片区"), region=cell(r, "大区"),
                city=cell(r, "市"), addr=cell(r, "进店地址"),
                ctype=cell(r, "客户类型", "门店"),
                auto=cell(r, "自动离店", "否"), chk=cell(r, "打卡状态"),
                ou=cell(r, "办事处"),
            ))
        stats = CleaningStats(raw_events=len(raw))

        # ---- R3: GPS 偏差 (独立标记) ----
        flags: List[EventFlag] = []
        for e in raw:
            e["gps_ok"] = (e["dev"] <= params.r3_deviation_max_m
                           and e["lon"] != 0.0 and e["lat"] != 0.0)
            if not e["gps_ok"]:
                stats.r3_gps_bad += 1
                flags.append(EventFlag("R3_GPS_DEVIANCE", f"EVT:{e['code']}@{e['t_in']}",
                                       f"dev={e['dev']:.0f}m"))

        # ---- R1 + R2 (按 rep+day 时序重排) ----
        by_rep_day: Dict[Tuple[str, datetime.date], List[dict]] = defaultdict(list)
        for e in raw:
            by_rep_day[(e["rep"], e["t_in"].date())].append(e)

        for key in sorted(by_rep_day):
            evs = sorted(by_rep_day[key], key=lambda e: e["t_in"])
            prev_exit = None
            prev_coord = None
            for e in evs:
                e["dur_clean"] = e["dur"]
                if e["dur"] >= params.r1_cap_min or e["auto"] == "是":
                    e["dur_clean"] = float(params.r1_cap_min)
                    stats.r1_truncated += 1
                    flags.append(EventFlag(
                        "R1_CHECKOUT_INFLATION", f"EVT:{e['code']}@{e['t_in']}",
                        f"dur={e['dur']:.0f}min->{e['dur_clean']:.0f}min(auto={e['auto']})"))
                e["r2"] = False
                if prev_exit is not None and prev_coord is not None:
                    gap = (e["t_in"] - prev_exit).total_seconds() / 60.0
                    same = (abs(e["lon"] - prev_coord[0]) <= params.r2_same_coord_eps
                            and abs(e["lat"] - prev_coord[1]) <= params.r2_same_coord_eps)
                    if same and gap <= params.r2_gap_max_min and e["dur"] <= params.r2_short_dur_max:
                        e["r2"] = True
                        stats.r2_suspects += 1
                        flags.append(EventFlag(
                            "R2_BATCH_CHECKIN_SUSPECT", f"EVT:{e['code']}@{e['t_in']}",
                            f"gap={gap:.1f}min,dur={e['dur']:.0f}min,same_coord"))
                e["credit"] = 0.0 if (e["r2"] or not e["gps_ok"]) else 1.0
                prev_exit = e["t_out"]
                prev_coord = (e["lon"], e["lat"])

        stats.effective_events = sum(1 for e in raw if e["credit"] > 0)

        # ---- R4: 客户坐标漂移 -> 聚类质心 ----
        cust_coords: Dict[str, List[GeoCoordinate]] = defaultdict(list)
        for e in raw:
            if e["gps_ok"]:
                cust_coords[e["code"]].append(GeoCoordinate(round(e["lon"], 6), round(e["lat"], 6)))

        cust_loc: Dict[str, Tuple[Optional[GeoCoordinate], GeoQualityStatus, int]] = {}
        last_by_cust: Dict[str, dict] = {}
        for e in raw:
            last_by_cust[e["code"]] = e

        for code, coords in cust_coords.items():
            stats.r4_total_customers += 1
            span = max(max(c.longitude for c in coords) - min(c.longitude for c in coords),
                       max(c.latitude for c in coords) - min(c.latitude for c in coords))
            if span > params.r4_drift_deg:
                stats.r4_drift_customers += 1
                members, n = _coord_cluster(coords, params.r4_drift_deg)
                if members:
                    cen = GeoCoordinate(round(mean(m.longitude for m in members), 6),
                                        round(mean(m.latitude for m in members), 6))
                    # canonical 枚举仅 EXACT_MATCH/UNMAPPED; 漂移店以簇质心保留可规划性,
                    # 漂移证据写入 R4 flag (不污染 DTO)
                    cust_loc[code] = (cen, GeoQualityStatus.EXACT_MATCH, n)
                    flags.append(EventFlag("R4_STORE_COORD_DRIFT", f"CUST:{code}",
                                           f"span={span:.4f}deg,cluster={n}/{len(coords)}"))
                    continue
            cust_loc[code] = (coords[0], GeoQualityStatus.EXACT_MATCH, len(coords))
        for code in last_by_cust:
            if code not in cust_loc:
                cust_loc[code] = (None, GeoQualityStatus.UNMAPPED, 0)
                stats.unmapped_customers += 1

        # ---- 归属: 每客户主覆盖 rep (事件数多者) ----
        rep_count_by_cust: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        rep_assigned: Dict[str, Set[str]] = defaultdict(set)
        for e in raw:
            rep_count_by_cust[e["code"]][e["rep"]] += 1
            rep_assigned[e["rep"]].add(e["code"])
        ownership = {code: max(counts.items(), key=lambda kv: kv[1])[0]
                     for code, counts in rep_count_by_cust.items()}

        # ---- 有效月度频次观测 -> 合成 OperationalVisitPolicy ----
        eff_month_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for e in raw:
            if e["credit"] > 0:
                eff_month_counts[e["code"]][e["t_in"].strftime("%Y-%m")] += 1

        customers: Dict[str, OperationalCustomer] = {}
        policies: Dict[str, OperationalVisitPolicy] = {}
        for code, e in last_by_cust.items():
            loc, geo_q, _n = cust_loc[code]
            customers[code] = OperationalCustomer(
                store_code=code,
                store_name=e["name"],
                tier="STANDARD",
                ka_name=e["ctype"],
                district=e["org"] or "未知片区",
                location=loc,
                geo_quality=geo_q,
                fulfillment_class=FulfillmentClass.COMMITTED,
                planned_frequency=None,  # canonical: source = PolicyRegistry
                address=e["addr"] or None,
            )
            per_month = eff_month_counts.get(code, {})
            avg_f = max(1, min(4, round(mean(per_month.values())))) if per_month else 1
            policies[code] = OperationalVisitPolicy(
                policy_id=f"OBS_FREQ_{code}",
                policy_version="sfa-obs-v1.0",
                store_code=code,
                target_frequency_per_month=avg_f,
                cadence_type="OBSERVED_FLEXIBLE",
                same_weekday_locked=False,
                bitemporal=BitemporalPeriod(
                    valid_from=datetime.datetime(2026, 1, 1, tzinfo=TZ_CN),
                    valid_to=datetime.datetime(2027, 12, 31, tzinfo=TZ_CN),
                    transaction_from=assembled_at),
                approved_by="DERIVED_FROM_OBSERVATION",  # 派生标注, 非签署
            )

        # 渠道层级 (客户类型 -> AccountHierarchyEntity)
        hierarchies: Dict[str, AccountHierarchyEntity] = {}
        for t in sorted({c.ka_name for c in customers.values()}):
            hid = f"CH_{hashlib.sha256(t.encode('utf-8')).hexdigest()[:6]}"
            hierarchies[hid] = AccountHierarchyEntity(hid, t, ChannelTier.TRADITIONAL)
            for code, c in customers.items():
                if c.ka_name == t:
                    customers[code] = replace(c, account_hierarchy_ref=hid)

        registry = PolicyRegistry(
            cadence_rules={
                "RULE_STRICT_WEEKLY": CadenceRule("RULE_STRICT_WEEKLY", 4, "STRICT_WEEKLY", 7, True),
                "RULE_STRICT_BIWEEKLY": CadenceRule("RULE_STRICT_BIWEEKLY", 2, "STRICT_BIWEEKLY", 14, True),
                "RULE_STRICT_MONTHLY": CadenceRule("RULE_STRICT_MONTHLY", 1, "STRICT_MONTHLY", 28, True),
            },
            ownership_map=ownership,
            operational_policies=policies,
        )

        # ---- resources: depot = 归属店可信坐标质心 ----
        resources: Dict[str, OperationalResource] = {}
        rep_coords: Dict[str, List[GeoCoordinate]] = defaultdict(list)
        for code, owner in ownership.items():
            loc, _gq, _ = cust_loc[code]
            if loc is not None:
                rep_coords[owner].append(loc)
        rep_meta: Dict[str, dict] = {}
        for e in raw:
            rep_meta.setdefault(e["rep"], dict(region=e["region"], sub_region=e["ou"], city=e["city"]))
        for rep, codes in rep_assigned.items():
            coords = rep_coords.get(rep, [])
            centroid = (GeoCoordinate(round(mean(c.longitude for c in coords), 6),
                                      round(mean(c.latitude for c in coords), 6))
                        if coords else GeoCoordinate(113.27, 23.10))
            meta = rep_meta.get(rep, {})
            resources[rep] = OperationalResource(
                rep_id=rep, rep_name=rep,
                region=meta.get("region", "未知大区"),
                sub_region=meta.get("sub_region", "未知办事处"),
                city=meta.get("city", "未知市"),
                depot_estimate=DerivedDepotEstimate(
                    rep_id=rep, inferred_centroid=centroid,
                    sample_points_count=len(coords),
                    confidence_score=round(min(0.95, 0.5 + 0.01 * len(coords)), 2) if coords else 0.3),
                assigned_store_codes=tuple(sorted(codes)),
                max_daily_stops=6, max_daily_workload_min=480.0,
            )

        # ---- 清洗后执行事实流 ----
        events: List[ActualVisitEvent] = []
        for e in sorted(raw, key=lambda x: (x["rep"], x["t_in"])):
            notes = []
            if e.get("dur_clean") != e["dur"]:
                notes.append(f"R1:dur={e['dur']:.0f}->{e['dur_clean']:.0f}")
            if e["r2"]:
                notes.append("R2:batch_suspect")
            if not e["gps_ok"]:
                notes.append("R3:gps_dev")
            actions = ()
            if notes:
                actions = (InStoreActionFact(
                    InStoreActionType.OUT_OF_STOCK_REMEDY, 0.0,
                    ";".join(notes) + f";credit={e['credit']:.1f}"),)
            events.append(ActualVisitEvent(
                event_id=f"SFA_{e['row']:06d}",
                store_code=e["code"], rep_id=e["rep"],
                visit_date=e["t_in"].date(),
                service_duration_min=e["dur_clean"],
                transit_duration_min=0.0,
                is_line_internal=(e["chk"] == "正常"),
                actions=actions,
                summary="|".join(notes),
            ))

        dates = [e["t_in"] for e in raw]
        bitemporal = BitemporalPeriod(
            valid_from=min(dates).replace(tzinfo=TZ_CN),
            valid_to=(max(dates) + datetime.timedelta(days=1)).replace(tzinfo=TZ_CN),
            transaction_from=assembled_at,
        )
        manifest = SourceManifest(
            source_file_path=str(path),
            source_file_sha256=sha256,
            raw_rows_count=len(rows) - 1,
            valid_facts_count=stats.effective_events,
            excluded_rows_count=excluded,
            exclusion_reason="rows missing 客户编码/人员名称/时间字段",
            assembled_at=assembled_at,
            loader_version="SFACheckinIngestor_v1.0",
        )

        ws = OperationalDecisionWorldState(
            snapshot_id=snapshot_id,
            bitemporal=bitemporal,
            manifest=manifest,
            customers=customers,
            resources=resources,
            account_hierarchies=hierarchies,
            product_line_scopes={},
            supply_nodes={},
            policies=registry,
            execution_fact_stream=events,
        )
        return ws, flags, stats
